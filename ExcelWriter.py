import os
import openpyxl
from education import UniversityService, Faculty, Student
from openpyxl.styles import Alignment

#Запис у ексель-таблиці
class ExcelWriter:
    def __init__(self, university_service: UniversityService):
        self.university_service = university_service

    def write_to_excel(self) -> None:
        filename = "University.xlsx"

        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(filename)

        universities = self.university_service.get_all()
        all_students = {}  # Используем словарь для хранения студентов и их баллов

        for university_name, university in universities.items():
            for faculty_name, faculty in university.faculties.items():
                sheet_name = f"{university_name[:10]}-{faculty_name[:10]}-{faculty.program[:10]}"
                sheet_name = sheet_name[:31]
                original_sheet_name = sheet_name
                suffix = 1
                while sheet_name in workbook.sheetnames:
                    sheet_name = f"{original_sheet_name[:28]}_{suffix}"
                    suffix += 1

                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet.delete_rows(1, sheet.max_row)
                else:
                    sheet = workbook.create_sheet(title=sheet_name)

                row = 1
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                sheet.cell(row=row, column=1, value=f"Университет: {university_name}")
                row += 1

                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                sheet.cell(row=row, column=1, value=f"Факультет: {faculty_name}")
                row += 1

                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                sheet.cell(row=row, column=1, value=f"Программа: {faculty.program}")
                row += 1

                # Заголовки для студентов
                sheet.cell(row=row, column=1, value="№")
                sheet.cell(row=row, column=2, value="ФИО")
                sheet.cell(row=row, column=3, value="Бал")
                row += 1

                # Запись студентов
                for index, student in enumerate(faculty.students, start=1):
                    sheet.cell(row=row, column=1, value=index)
                    sheet.cell(row=row, column=2, value=student.name)
                    sheet.cell(row=row, column=3, value=student.rating)
                    row += 1
                    all_students[student.name] = student.rating  # Добавление студента и его балла в словарь

        # Создание сводного листа для заявок
        summary_sheet = workbook.create_sheet(title="Applications")
        col = 3  # Столбец для университетов, факультетов и программ
        faculties_info = []

        # Заголовок для университетов
        for university_name, university in universities.items():
            for faculty_name, faculty in university.faculties.items():
                # Название университета в первом рядке
                summary_sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col)
                summary_sheet.cell(row=1, column=col, value=university_name).alignment = Alignment(horizontal="center")
                
                # Название факультета во втором рядке
                summary_sheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col)
                summary_sheet.cell(row=2, column=col, value=faculty_name).alignment = Alignment(horizontal="center")
                
                # Название программы в третьем рядке
                summary_sheet.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col)
                summary_sheet.cell(row=3, column=col, value=faculty.program).alignment = Alignment(horizontal="center")

                # Сохранение факультетов для дальнейшего использования
                faculties_info.append((university_name, faculty_name, faculty))
                
                # Переход к следующему столбцу для следующего факультета
                col += 1

        # Заголовки для ФИО и Баллов
        summary_sheet.cell(row=3, column=1, value="ФИО")
        summary_sheet.cell(row=3, column=2, value="Балл")

        # Запись всех уникальных студентов и их баллов начиная с 4-го ряда
        all_students = dict(sorted(all_students.items()))  # Сортировка студентов для упорядоченного отображения
        row = 4
        for student_name, student_rating in all_students.items():
            summary_sheet.cell(row=row, column=1, value=student_name)
            summary_sheet.cell(row=row, column=2, value=student_rating)
            row += 1

        # Заполнение плюсов и минусов
        for i, student_name in enumerate(all_students, start=4):
            for j, (_, _, faculty) in enumerate(faculties_info, start=3):
                if any(student.name == student_name for student in faculty.students):
                    summary_sheet.cell(row=i, column=j, value="+").alignment = Alignment(horizontal="center")
                else:
                    summary_sheet.cell(row=i, column=j, value="-").alignment = Alignment(horizontal="center")

        # Автоматическая подгонка ширины столбцов
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            summary_sheet.column_dimensions[column_letter].width = adjusted_width

        # Удаляем пустой стандартный лист
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

        # Сохраняем книгу
        workbook.save(filename)